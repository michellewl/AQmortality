# Imports

import numpy as np
import requests
import pandas as pd
from os import makedirs, path, listdir
from tqdm import tqdm
import zipfile as zpf
import matplotlib.pyplot as plt
import xlrd
from openpyxl import load_workbook
import wandb

# London Air Quality Network data class

class LAQNData():
    def __init__(self, url, species, start_date, end_date):
        self.url = url
        self.species = species
        self.start_date = start_date
        self.end_date = end_date
        
        london_sites = requests.get(self.url)
        self.sites_df = pd.DataFrame(london_sites.json()['Sites']['Site'])
        self.site_codes = self.sites_df["@SiteCode"].tolist()

    def download(self, verbose=True):
        laqn_df = pd.DataFrame()
        
        if verbose:
            progress_bar = tqdm(self.site_codes)
        else:
            progress_bar = self.site_codes
            
        for site_code in progress_bar:
            if verbose:
                progress_bar.set_description(f'Working on site {site_code}')
            url_species = f"http://api.erg.kcl.ac.uk/AirQuality/Data/SiteSpecies/SiteCode={site_code}/SpeciesCode={self.species}/StartDate={self.start_date}/EndDate={self.end_date}/csv"
            cur_df = pd.read_csv(url_species)
            cur_df.columns = ["date", site_code]
            cur_df.set_index("date", drop=True, inplace=True)

            try:
                if laqn_df.empty:
                    laqn_df = cur_df.copy()
                else:
                    laqn_df = laqn_df.join(cur_df.copy(), how="outer")

            except ValueError:  # Trying to join with duplicate column names
                rename_dict = {}
                for x in list(set(cur_df.columns).intersection(laqn_df.columns)):
                    rename_dict.update({x: f"{x}_"})
                    print(f"Renamed duplicated column:\n{rename_dict}")
                laqn_df.rename(mapper=rename_dict, axis="columns", inplace=True)
                if laqn_df.empty:
                    laqn_df = cur_df.copy()
                else:
                    laqn_df = laqn_df.join(cur_df.copy(), how="outer")
                if verbose:
                    print(f"Joined.")

            except KeyError:  # Trying to join along indexes that don't match
                print(f"Troubleshooting {site_code}...")
                cur_df.index = cur_df.index + ":00"
                if laqn_df.empty:
                    laqn_df = cur_df.copy()
                else:
                    laqn_df = laqn_df.join(cur_df.copy(), how="outer")
                print(f"{site_code} joined.")

        #print("Data download complete. Removing sites with 0 data...")
        laqn_df.dropna(axis="columns", how="all", inplace=True)
        return laqn_df
        
    def download_and_log(self):
        with wandb.init(project="AQmortality", job_type="load-data") as run:
            df = self.download()
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "laqn-raw", type="dataset",
                description=f"Raw LAQN {self.species} data from {self.start_date} to {self.end_date}, split according to site codes.",
                metadata={"source":self.url,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)


    
    def read(self, sites, artifact):
        with wandb.init(project="AQmortality", job_type="read-data") as run:
            raw_data_artifact = run.use_artifact(f'{artifact}:latest')
            data_folder = raw_data_artifact.download()
            df = pd.DataFrame()
            empty_sites = []
            for site in sites:
                filepath = path.join(data_folder, f"{site}.npz")
                try:
                    data = np.load(filepath, allow_pickle=True)
                except FileNotFoundError:
                    empty_sites.append(site)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))
            empty_sites = ", ".join(empty_sites)
            print(f"No data for site codes: {empty_sites}")
        return df
    

        
    def resample_time_and_log(self, sites, date_index):
        
        with wandb.init(project="AQmortality", job_type="resample-data") as run:
            raw_data_artifact = run.use_artifact('laqn-raw:latest')
            data_folder = raw_data_artifact.download()
            df = pd.DataFrame()
            for site in sites:
                filepath = path.join(data_folder, f"{site}.npz")
                try:
                    data = np.load(filepath, allow_pickle=True)
                except FileNotFoundError:
                    continue
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))

            df = df.loc[df.index < date_index.max()]
            df = df.loc[df.index > date_index.min()]
            resampled_df = df.groupby(date_index[date_index.searchsorted(df.index)]).mean()
            columns = resampled_df.columns.to_list()
            resample_data = wandb.Artifact(
                "laqn-resample", type="dataset",
                description=f"Resampled LAQN {self.species} data from {self.start_date} to {self.end_date}, split according to site codes.",
                metadata={"source":self.url,
                         "shapes":[resampled_df[column].shape for column in columns],
                         "sites":columns,
                        "species":self.species,
                        "start_date":self.start_date,
                        "end_date":self.end_date})
            for column in columns:
                with resample_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=resampled_df.index, y=resampled_df[column].values)

            run.log_artifact(resample_data)
        
        return resampled_df  
    
    def regional_average_and_log(self, sites):
        with wandb.init(project="AQmortality", job_type="regional-average-data") as run:
            daily_data_artifact = run.use_artifact('laqn-resample:latest')
            data_folder = daily_data_artifact.download()
            df = pd.DataFrame()
            for site in sites:
                filepath = path.join(data_folder, f"{site}.npz")
                try:
                    data = np.load(filepath, allow_pickle=True)
                except FileNotFoundError:
                    continue
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))

            df = pd.DataFrame(df.mean(axis=1), columns=["mean_NO2"])
            columns = df.columns.to_list()
            regional_data = wandb.Artifact(
                "laqn-regional", type="dataset",
                description=f"Regional average LAQN {self.species} data from {self.start_date} to {self.end_date}.",
                metadata={"source":self.url,
                         "shapes":[df[column].shape for column in columns],
                         "species":self.species,
                        "start_date":self.start_date,
                        "end_date":self.end_date})
            for column in columns:
                with regional_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(regional_data)
        
        return df

    
# Office for National Statistics health data class

class HealthData():
    def __init__(self):
        self.tmp_folder = path.join(path.abspath(""), "tmp")
        
        if not path.exists(self.tmp_folder):
            makedirs(self.tmp_folder)

    def download(self, url, verbose=False):
        self.filename = path.basename(url)
        _, self.extension = path.splitext(self.filename)
        self.filepath = path.join(self.tmp_folder, self.filename)
        
        request = requests.get(url)
        file = open(self.filepath, 'wb')
        file.write(request.content)
        file.close()
        if verbose:
            print(f"Saved to {self.filename}")
        if self.extension == ".zip":
            self.zipfiles = zpf.ZipFile(self.filepath).namelist()
            if verbose:
                print("Contains zip files:")
                [print(f"[{i}] {self.zipfiles[i]}") for i in range(len(self.zipfiles))]
        elif self.extension == ".xls":
            workbook = xlrd.open_workbook(self.filepath)
            self.sheets = workbook.sheet_names()
            if verbose:
                print(f"Contains xls sheets: {self.sheets}")
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            self.sheets = workbook.sheetnames
            if verbose:
                print(f"Contains xlsx sheets: {self.sheets}")
                
    def unzip(self, file="all", verbose=False):
        with zpf.ZipFile(self.filepath, 'r') as zip_ref:
            if file =="all":                                       # Extract all zipped files.
                zip_ref.extractall(self.tmp_folder)
            else:                                                  # Extract one specified file,
                zip_ref.extract(file, self.tmp_folder)            # then
                self.filename = path.basename(file)                # reset the file name, path and extension info.
                _, self.extension = path.splitext(self.filename)
                self.filepath = path.join(self.tmp_folder, self.filename)
             
            if verbose:
                print(f"Unzipped {file}.")
            if self.extension == ".xls":
                workbook = xlrd.open_workbook(self.filepath)
                self.sheets = workbook.sheet_names()
                if verbose:
                    print(f"Contains xls sheets: {self.sheets}")
            elif self.extension == ".xlsx":
                workbook = load_workbook(self.filepath)
                self.sheets = workbook.sheetnames
                if verbose:
                    print(f"Contains xlsx sheets: {self.sheets}")
        
    def download_and_log(self, region_code, start_year, end_year, url_dict):
        with wandb.init(project="AQmortality", job_type="load-data") as run:
            df = pd.DataFrame()

            for year in range(start_year, end_year+1):
                url = url_dict[year]

                if year == start_year or not url == urls[year-1]:
                    self.download(url, verbose=False)
                    if self.extension == ".zip":
                        self.unzip(self.zipfiles[0], verbose=False)
                    adhoc_df = self.read_xls(self.sheets[-1], verbose=False)
                    adhoc_df.columns = range(adhoc_df.shape[1])
                    adhoc_df = adhoc_df.loc[adhoc_df[3] == region_code, [0, 1, 2, 3, 4]]
                    adhoc_df.columns = ["year", "month", "day", "region_code", "deaths"]
                    adhoc_df["date"] = pd.to_datetime(adhoc_df[["year", "month", "day"]])

                df = df.append(adhoc_df.loc[adhoc_df["year"] == year, ["date", "deaths"]].copy().set_index("date"))
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "mortality-raw", type="dataset",
                description=f"Raw daily mortality data for total region {region_code}. Data is extracted from source Excel files (not logged using Weights and Biases).",
                metadata={"source":url_dict,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)
        
        [remove(path.join(self.tmp_folder, file)) for file in listdir(self.tmp_folder)]
        
    def read_csv(self, verbose=True, index_col="date", parse_dates=True):
        if verbose:
            print(f"Reading {self.filename}...")
        return pd.read_csv(self.filepath, index_col=index_col, parse_dates=parse_dates)
    
    def read_xls(self, sheet_name, verbose=False):
        if verbose:
            print(f"Reading {self.filename}...")
        if self.extension == ".xls":
            return pd.read_excel(self.filepath, sheet_name)
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            worksheet = workbook[sheet_name]
            return pd.DataFrame(worksheet.values)
        
    def read(self, artifact):
        with wandb.init(project="AQmortality", job_type="read-data") as run:
            raw_data_artifact = run.use_artifact(f"{artifact}:latest")
            data_folder = raw_data_artifact.download()
            filepath = path.join(data_folder, f"deaths.npz")
            data = np.load(filepath, allow_pickle=True)
            df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=["deaths"])
        return df
    
    def scale_per_capita_and_log(self):
        with wandb.init(project="AQmortality", job_type="scale-data") as run:
            mort_data_artifact = run.use_artifact('mortality-raw:latest')
            data_folder = mort_data_artifact.download()
            filepath = path.join(data_folder, f"deaths.npz")
            mort_data = np.load(filepath, allow_pickle=True)
            mort_df = pd.DataFrame(index=pd.DatetimeIndex(mort_data["x"]), data=mort_data["y"].astype(int), columns=["deaths"])
            
            pop_data_artifact = run.use_artifact('population-resample:latest')
            data_folder = pop_data_artifact.download()
            filepath = path.join(data_folder, f"population.npz")
            pop_data = np.load(filepath, allow_pickle=True)
            pop_df = pd.DataFrame(index=pd.DatetimeIndex(pop_data["x"]), data=pop_data["y"].astype(int), columns=["population"])
                           
            df = pop_df.join(mort_df).dropna()
            df["deaths_per_capita"] = df["deaths"]/df["population"]
            
            columns = df.columns.to_list()
            scale_data = wandb.Artifact(
                "mortality-scaled", type="dataset",
                description=f"Scaled mortality data per capita for total London region.",
                metadata={"shapes":[df[column].shape for column in columns],
                         "columns":columns})
            for column in columns:
                with scale_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(scale_data)
        
        return df
            
        
# Meteorology data class

class MetData():
    def __init__(self, station):
        self.station = station
            
    def download(self, url):
        columns = ["date", "hour", "temperature", "dew_point", "humidity", "precip", "blank1", "wind_dir", "wind_speed", "peak_gust", "pressure", "blank2", "blank3"]
        df = pd.read_csv(url, header=None, names=columns).drop(["blank1", "blank2", "blank3"], axis=1)
        df["date"] = df["date"] + " " + df["hour"].astype(str) +":00"
        df = df.drop(["hour"], axis=1).set_index("date")
        return df
        
    def download_and_log(self, url):
        with wandb.init(project="AQmortality", job_type="load-data") as run:
            df = self.download(url)
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "met-raw", type="dataset",
                description=f"Raw meteorology data from the {self.station} weather station, split according to meteorological variables.",
                metadata={"source":url,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)
    
    def read(self, variables, artifact):
        with wandb.init(project="AQmortality", job_type="read-data") as run:
            raw_data_artifact = run.use_artifact(f"{artifact}:latest")
            data_folder = raw_data_artifact.download()
            df = pd.DataFrame()
            for variable in variables:
                filepath = path.join(data_folder, f"{variable}.npz")
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable]))
        return df
    
    def resample_time_and_log(self, date_index):
        variables = ["temperature", "dew_point", "humidity", "precip", "wind_dir", "wind_speed", "peak_gust", "pressure"]
        
        with wandb.init(project="AQmortality", job_type="resample-data") as run:
            raw_data_artifact = run.use_artifact('met-raw:latest')
            data_folder = raw_data_artifact.download()
            df = pd.DataFrame()
            for variable in variables:
                filepath = path.join(data_folder, f"{variable}.npz")
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable]))

            df = df.loc[df.index < date_index.max()]
            df = df.loc[df.index > date_index.min()]
            resampled_df = df.groupby(date_index[date_index.searchsorted(df.index)]).mean()
            columns = resampled_df.columns.to_list()
            resample_data = wandb.Artifact(
                "met-resample", type="dataset",
                description=f"Resampled meteorology data from the {self.station} weather station, split according to meteorological variables.",
                metadata={"shapes":[resampled_df[column].shape for column in columns],
                         "columns":columns})
            for column in columns:
                with resample_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=resampled_df.index, y=resampled_df[column].values)

            run.log_artifact(resample_data)
        
        return resampled_df
    
    
# Population data class

class PopData():
    def __init__(self):
        self.tmp_folder = path.join(path.abspath(""), "tmp")
        
        if not path.exists(self.tmp_folder):
            makedirs(self.tmp_folder)

    def download(self, url, verbose=False):
        self.filename = path.basename(url)
        _, self.extension = path.splitext(self.filename)
        self.filepath = path.join(self.tmp_folder, self.filename)
        
        request = requests.get(url)
        file = open(self.filepath, 'wb')
        file.write(request.content)
        file.close()
        if verbose:
            print(f"Saved to {self.filename}")
        if self.extension == ".zip":
            self.zipfiles = zpf.ZipFile(self.filepath).namelist()
            if verbose:
                print("Contains zip files:")
                [print(f"[{i}] {self.zipfiles[i]}") for i in range(len(self.zipfiles))]
        elif self.extension == ".xls":
            workbook = xlrd.open_workbook(self.filepath)
            self.sheets = workbook.sheet_names()
            if verbose:
                print(f"Contains xls sheets: {self.sheets}")
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            self.sheets = workbook.sheetnames
            if verbose:
                print(f"Contains xlsx sheets: {self.sheets}")
                
    def unzip(self, file="all", verbose=False):
        with zpf.ZipFile(self.filepath, 'r') as zip_ref:
            if file =="all":                                       # Extract all zipped files.
                zip_ref.extractall(self.tmp_folder)
            else:                                                  # Extract one specified file,
                zip_ref.extract(file, self.tmp_folder)            # then
                self.filename = path.basename(file)                # reset the file name, path and extension info.
                _, self.extension = path.splitext(self.filename)
                self.filepath = path.join(self.tmp_folder, self.filename)
             
            if verbose:
                print(f"Unzipped {file}.")
            if self.extension == ".xls":
                workbook = xlrd.open_workbook(self.filepath)
                self.sheets = workbook.sheet_names()
                if verbose:
                    print(f"Contains xls sheets: {self.sheets}")
            elif self.extension == ".xlsx":
                workbook = load_workbook(self.filepath)
                self.sheets = workbook.sheetnames
                if verbose:
                    print(f"Contains xlsx sheets: {self.sheets}")
        
    def download_and_log(self, url, region_name):
        with wandb.init(project="AQmortality", job_type="load-data") as run:
            self.download(url)
    
            df = pd.DataFrame()

            for sheet in ["Table 3", "Table 4"]:
                tmp_df = self.read_xls(sheet, verbose=False)
                tmp_df = tmp_df.loc[tmp_df[0].apply(lambda x: isinstance(x, int))].reset_index(drop=True)[[0]].rename(columns={0:"date"}).join(tmp_df.loc[tmp_df[1]=="London"].reset_index(drop=True)[[2]].rename(columns={2:"population"})).set_index("date")
                tmp_df.index = pd.to_datetime(tmp_df.index, format="%Y") + pd.tseries.offsets.DateOffset(months=6) # Set these as mid-year estimates
                if df.empty:
                    df = tmp_df.copy()
                else:
                    df = df.append(tmp_df.copy())
                    
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "population-raw", type="dataset",
                description=f"Raw annual population data for total {region_name} region. Data is extracted from source Excel file (not logged using Weights and Biases).",
                metadata={"source":url,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)
        
        [remove(path.join(self.tmp_folder, file)) for file in listdir(self.tmp_folder)]
        
    def read_csv(self, verbose=True, index_col="date", parse_dates=True):
        if verbose:
            print(f"Reading {self.filename}...")
        return pd.read_csv(self.filepath, index_col=index_col, parse_dates=parse_dates)
    
    def read_xls(self, sheet_name, verbose=False):
        if verbose:
            print(f"Reading {self.filename}...")
        if self.extension == ".xls":
            return pd.read_excel(self.filepath, sheet_name)
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            worksheet = workbook[sheet_name]
            return pd.DataFrame(worksheet.values)
        
    def read(self, artifact):
        with wandb.init(project="AQmortality", job_type="read-data") as run:
            raw_data_artifact = run.use_artifact(f'{artifact}:latest')
            data_folder = raw_data_artifact.download()
            filepath = path.join(data_folder, f"population.npz")
            data = np.load(filepath, allow_pickle=True)
            df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=["population"])
        return df
    
    def resample_time_and_log(self, key, method):
        with wandb.init(project="AQmortality", job_type="resample-data") as run:
            raw_data_artifact = run.use_artifact('population-raw:latest')
            data_folder = raw_data_artifact.download()
            filepath = path.join(data_folder, f"population.npz")
            data = np.load(filepath, allow_pickle=True)
            df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"].astype(int), columns=["population"])
            
            resampled_df = df.resample(key).asfreq().interpolate(method=method)
            columns = resampled_df.columns.to_list()
            resample_data = wandb.Artifact(
                "population-resample", type="dataset",
                description=f"Resampled population data to daily resolution by {method} interpolation.",
                metadata={"shapes":[resampled_df[column].shape for column in columns],
                         "columns":columns,
                        "key":key,
                         "method":method})
            for column in columns:
                with resample_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=resampled_df.index, y=resampled_df[column].values)

            run.log_artifact(resample_data)
        
        return resampled_df  