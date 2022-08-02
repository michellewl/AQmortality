# Imports

import numpy as np
import requests
import pandas as pd
import geopandas as gpd
from os import makedirs, path, listdir, remove
from tqdm import tqdm
import zipfile as zpf
import matplotlib.pyplot as plt
import xlrd
from openpyxl import load_workbook
import wandb
import httplib2
from bs4 import BeautifulSoup, SoupStrainer
from shutil import rmtree

project = "AQmortality"

# London Air Quality Network data class

class LAQNData():
    def __init__(self, species, region):
        self.species = species
        self.region = region
        self.url = f"http://api.erg.kcl.ac.uk/AirQuality/Information/MonitoringSites/GroupName={self.region}/Json"
               
        london_sites = requests.get(self.url)
        self.sites_df = pd.DataFrame(london_sites.json()['Sites']['Site'])
        self.site_codes = self.sites_df["@SiteCode"].tolist()

    def download(self, start_date, end_date, verbose=True):
        self.start_date = start_date
        self.end_date = end_date
        laqn_df = pd.DataFrame()
        
        if verbose:
            progress_bar = tqdm(self.site_codes)
        else:
            progress_bar = self.site_codes
            
        for site_code in progress_bar:
            if verbose:
                progress_bar.set_description(f'Working on site {site_code}')
            url_species = f"http://api.erg.kcl.ac.uk/AirQuality/Data/SiteSpecies/SiteCode={site_code}/SpeciesCode={self.species}/StartDate={self.start_date}/EndDate={self.end_date}/csv"
            cur_df = pd.read_csv(url_species)
            cur_df.columns = ["date", site_code]
            cur_df.set_index("date", drop=True, inplace=True)

            try:
                if laqn_df.empty:
                    laqn_df = cur_df.copy()
                else:
                    laqn_df = laqn_df.join(cur_df.copy(), how="outer")

            except ValueError:  # Trying to join with duplicate column names
                rename_dict = {}
                for x in list(set(cur_df.columns).intersection(laqn_df.columns)):
                    rename_dict.update({x: f"{x}_"})
                    print(f"Renamed duplicated column:\n{rename_dict}")
                laqn_df.rename(mapper=rename_dict, axis="columns", inplace=True)
                if laqn_df.empty:
                    laqn_df = cur_df.copy()
                else:
                    laqn_df = laqn_df.join(cur_df.copy(), how="outer")
                if verbose:
                    print(f"Joined.")

            except KeyError:  # Trying to join along indexes that don't match
                print(f"Troubleshooting {site_code}...")
                cur_df.index = cur_df.index + ":00"
                if laqn_df.empty:
                    laqn_df = cur_df.copy()
                else:
                    laqn_df = laqn_df.join(cur_df.copy(), how="outer")
                print(f"{site_code} joined.")

        #print("Data download complete. Removing sites with 0 data...")
        laqn_df.dropna(axis="columns", how="all", inplace=True)
        return laqn_df
        
    def download_and_log(self, start_date, end_date):
        with wandb.init(project=project, job_type="load-data") as run:
            df = self.download(start_date, end_date)
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "laqn-raw", type="dataset",
                description=f"Raw LAQN {self.species} data for the {self.region} region from {self.start_date} to {self.end_date}, split according to site codes.",
                metadata={"source":self.url,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)


    
    def read(self, artifact_name):
        artifact = wandb.Api().artifact(f"{project}/{artifact_name}:latest")
        data_folder = artifact.download()
        df = pd.DataFrame()
        if artifact_name == "laqn-regional":
            filepath = path.join(data_folder, f"mean_{self.species}.npz")
            data = np.load(filepath, allow_pickle=True)
            df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[f"mean_{self.species}"])
        else:
            for file in listdir(data_folder):
                site = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))
                               
        return df
    

        
    def resample_time_and_log(self, date_index):
        
        with wandb.init(project=project, job_type="resample-data") as run:
            artifact = run.use_artifact('laqn-raw:latest')
            data_folder = artifact.download()
            df = pd.DataFrame()
            for file in listdir(data_folder):
                site = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))

            df = df.loc[df.index < date_index.max()]
            df = df.loc[df.index > date_index.min()]
            resampled_df = df.groupby(date_index[date_index.searchsorted(df.index)]).mean()
            columns = resampled_df.columns.to_list()
            resample_data = wandb.Artifact(
                "laqn-resample", type="dataset",
                description=f"Resampled LAQN {self.species} data for the {self.region} region from {df.index.min()} to {df.index.max()}, split according to site codes.",
                metadata={"source":self.url,
                         "shapes":[resampled_df[column].shape for column in columns],
                         "sites":columns,
                        "species":self.species})
            for column in columns:
                with resample_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=resampled_df.index, y=resampled_df[column].values)

            run.log_artifact(resample_data)
        
        return resampled_df  
    
    def regional_average_and_log(self):
        with wandb.init(project=project, job_type="regional-average-data") as run:
            artifact = run.use_artifact('laqn-resample:latest')
            data_folder = artifact.download()
            df = pd.DataFrame()
            for file in listdir(data_folder):
                site = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))

            df = pd.DataFrame(df.mean(axis=1), columns=[f"mean_{self.species}"])
            columns = df.columns.to_list()
            regional_data = wandb.Artifact(
                "laqn-regional", type="dataset",
                description=f"Regional {self.region} average LAQN {self.species} data from {df.index.min()} to {df.index.max()}.",
                metadata={"source":self.url,
                         "shapes":[df[column].shape for column in columns],
                         "species":self.species})
            for column in columns:
                with regional_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(regional_data)
        
        return df
    
    
    def local_authority_aggregation(self, df, verbose=True):
        print("Averaging sites according to local authorities...")
        meta_url = "http://api.erg.kcl.ac.uk/AirQuality/Information/MonitoringSites/GroupName=London/Json"
        sites_request = requests.get(meta_url)
        meta_df = pd.DataFrame(sites_request.json()['Sites']['Site'])

        df.index = pd.to_datetime(df.index)

        aggregated_df = pd.DataFrame()
        local_authorities_list = meta_df["@LocalAuthorityName"].unique().tolist()

        for i in tqdm(range(len(local_authorities_list))):
            local_authority = local_authorities_list[i]
            site_codes = meta_df.loc[meta_df["@LocalAuthorityName"] == local_authority, "@SiteCode"].values
            site_codes = list(set(site_codes).intersection(df.columns))
            current_df = df[site_codes].copy()

            if len(site_codes) == 0:
                if verbose:
                    print(f"No data for {local_authority}")
                continue

            if len(site_codes) == 1:
                current_df.columns = [local_authority]
                if verbose:
                    print(f"Skipped algorithm for {local_authority} - only one site with data.")
                if aggregated_df.empty:
                    aggregated_df = current_df.copy()
                else:
                    aggregated_df = aggregated_df.join(current_df.copy(), how="left")

            else:
                # Step 1: Compute annual mean for each monitor for each year
                annual_mean_df = current_df.resample("A").mean()

                # Step 2: Subtract annual mean from hourly measurements to obtain hourly deviance for the monitor
                for year in annual_mean_df.index.year:
                    for site in current_df.columns:
                        # Create a list of the annual mean value for the site that is the same length as the data
                        annual_mean = annual_mean_df.loc[annual_mean_df.index.year == year, site].tolist() * len(current_df.loc[current_df.index.year == year, site])
                        # Subtract the annual mean list from the dataframe
                        current_df.loc[current_df.index.year == year, site] = current_df.loc[current_df.index.year == year, site] - annual_mean
                # Calculate the annual mean for the borough (over multiple sites)
                annual_mean_df[local_authority] = annual_mean_df.mean(axis=1)

                # Step 3: Standardise the hourly deviance by dividing by standard deviation for the monitor
                sd_per_site = current_df.copy().std(axis=0, ddof=0)
                sd_per_borough = current_df.values.flatten()[~np.isnan(current_df.values.flatten())].std(ddof=0)
                current_df = current_df / sd_per_site

                # Step 4: Average the hourly standardised deviations to get an average across all monitors
                current_df[local_authority] = current_df.mean(axis=1)

                # Step 5: Multiply the hourly averaged standardised deviation
                # by the standard deviation across all monitor readings for the entire years (to un-standardise)
                current_df[local_authority] = current_df[local_authority] * sd_per_borough

                # Step 6: Add the hourly average deviance and annual average across all monitors to get a hourly average reading
                for year in annual_mean_df.index.year:
                    # Make the list the correct length as before
                    annual_mean = annual_mean_df.loc[annual_mean_df.index.year == year, local_authority].tolist() * len(current_df.loc[current_df.index.year == year])
                    # Add the annual mean
                    current_df.loc[current_df.index.year == year, local_authority] = current_df.loc[current_df.index.year == year, local_authority] + annual_mean
                # Only keep the hourly data for the borough
                current_df = current_df[[local_authority]]

                if aggregated_df.empty:
                    aggregated_df = current_df.copy()
                else:
                    aggregated_df = aggregated_df.join(current_df.copy(), how="left")

        return aggregated_df
    
    def local_authority_aggregation_and_log(self):
        with wandb.init(project=project, job_type="local-authority-average-data") as run:
            artifact = run.use_artifact('laqn-resample:latest')
            data_folder = artifact.download()
            df = pd.DataFrame()
            for file in listdir(data_folder):
                site = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))

            df = self.local_authority_aggregation(df)
            
            columns = df.columns.to_list()
            local_authority_data = wandb.Artifact(
                "laqn-local-authority", type="dataset",
                description=f"{self.region} local authority-aggregated LAQN {self.species} data from {df.index.min()} to {df.index.max()}.",
                metadata={"source":self.url,
                         "shapes":[df[column].shape for column in columns],
                         "species":self.species})
            for column in columns:
                with local_authority_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(local_authority_data)
        
        return df
    
    
# Office for National Statistics health data class

class HealthData():
    def __init__(self):
        self.tmp_folder = path.join(path.abspath(""), "tmp")
        
        if not path.exists(self.tmp_folder):
            makedirs(self.tmp_folder)

    def download(self, url, verbose=False):
        self.filename = path.basename(url)
        _, self.extension = path.splitext(self.filename)
        self.filepath = path.join(self.tmp_folder, self.filename)
        
        request = requests.get(url)
        file = open(self.filepath, 'wb')
        file.write(request.content)
        file.close()
        if verbose:
            print(f"Saved to {self.filename}")
        if self.extension == ".zip":
            self.zipfiles = zpf.ZipFile(self.filepath).namelist()
            if verbose:
                print("Contains zip files:")
                [print(f"[{i}] {self.zipfiles[i]}") for i in range(len(self.zipfiles))]
        elif self.extension == ".xls":
            workbook = xlrd.open_workbook(self.filepath)
            self.sheets = workbook.sheet_names()
            if verbose:
                print(f"Contains xls sheets: {self.sheets}")
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            self.sheets = workbook.sheetnames
            if verbose:
                print(f"Contains xlsx sheets: {self.sheets}")
                
    def unzip(self, file="all", verbose=False):
        with zpf.ZipFile(self.filepath, 'r') as zip_ref:
            if file =="all":                                       # Extract all zipped files.
                zip_ref.extractall(self.tmp_folder)
            else:                                                  # Extract one specified file,
                zip_ref.extract(file, self.tmp_folder)            # then
                self.filename = path.basename(file)                # reset the file name, path and extension info.
                _, self.extension = path.splitext(self.filename)
                self.filepath = path.join(self.tmp_folder, self.filename)
             
            if verbose:
                print(f"Unzipped {file}.")
            if self.extension == ".xls":
                workbook = xlrd.open_workbook(self.filepath)
                self.sheets = workbook.sheet_names()
                if verbose:
                    print(f"Contains xls sheets: {self.sheets}")
            elif self.extension == ".xlsx":
                workbook = load_workbook(self.filepath)
                self.sheets = workbook.sheetnames
                if verbose:
                    print(f"Contains xlsx sheets: {self.sheets}")
        
    def download_and_log(self, region_code, start_year, end_year, url_dict):
        with wandb.init(project=project, job_type="load-data") as run:
            df = pd.DataFrame()

            for year in range(start_year, end_year+1):
                url = url_dict[year]

                if year == start_year or not url == urls[year-1]:
                    self.download(url, verbose=False)
                    if self.extension == ".zip":
                        self.unzip(self.zipfiles[0], verbose=False)
                    adhoc_df = self.read_xls(self.sheets[-1], verbose=False)
                    adhoc_df.columns = range(adhoc_df.shape[1])
                    adhoc_df = adhoc_df.loc[adhoc_df[3] == region_code, [0, 1, 2, 3, 4]]
                    adhoc_df.columns = ["year", "month", "day", "region_code", "deaths"]
                    adhoc_df["date"] = pd.to_datetime(adhoc_df[["year", "month", "day"]])

                df = df.append(adhoc_df.loc[adhoc_df["year"] == year, ["date", "deaths"]].copy().set_index("date"))
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "mortality-raw", type="dataset",
                description=f"Raw daily mortality data for total region {region_code}. Data is extracted from source Excel files (not logged using Weights and Biases).",
                metadata={"source":url_dict,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)
        
        [remove(path.join(self.tmp_folder, file)) for file in listdir(self.tmp_folder)]
        
    def read_csv(self, verbose=True, index_col="date", parse_dates=True):
        if verbose:
            print(f"Reading {self.filename}...")
        return pd.read_csv(self.filepath, index_col=index_col, parse_dates=parse_dates)
    
    def read_xls(self, sheet_name, verbose=False):
        if verbose:
            print(f"Reading {self.filename}...")
        if self.extension == ".xls":
            return pd.read_excel(self.filepath, sheet_name)
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            worksheet = workbook[sheet_name]
            return pd.DataFrame(worksheet.values)
        
    def read(self, artifact_name):
        artifact = wandb.Api().artifact(f"{project}/{artifact_name}:latest")
        data_folder = artifact.download()
        filepath = path.join(data_folder, f"deaths.npz")
        data = np.load(filepath, allow_pickle=True)
        df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=["deaths"])
        return df
    
    def scale_per_capita_and_log(self):
        with wandb.init(project=project, job_type="scale-data") as run:
            artifact = run.use_artifact('mortality-raw:latest')
            data_folder = artifact.download()
            filepath = path.join(data_folder, f"deaths.npz")
            mort_data = np.load(filepath, allow_pickle=True)
            mort_df = pd.DataFrame(index=pd.DatetimeIndex(mort_data["x"]), data=mort_data["y"].astype(int), columns=["deaths"])
            
            artifact = run.use_artifact('population-resample:latest')
            data_folder = artifact.download()
            filepath = path.join(data_folder, f"population.npz")
            pop_data = np.load(filepath, allow_pickle=True)
            pop_df = pd.DataFrame(index=pd.DatetimeIndex(pop_data["x"]), data=pop_data["y"].astype(int), columns=["population"])
                           
            df = pop_df.join(mort_df).dropna()
            df["deaths_per_capita"] = df["deaths"]/df["population"]
            
            scale_data = wandb.Artifact(
                "mortality-scaled", type="dataset",
                description=f"Scaled mortality data per capita for total London region.",
                metadata={"shapes":df["deaths_per_capita"].shape,
                         "columns":"deaths"})
            with scale_data.new_file("deaths" + ".npz", mode="wb") as file:
                np.savez(file, x=df.index, y=df["deaths_per_capita"].values)

            run.log_artifact(scale_data)
        
        return df
            
        
# Meteorology data class

class MetData():
    def __init__(self, station):
        self.station = station
            
    def download(self, url):
        columns = ["date", "hour", "temperature", "dew_point", "humidity", "precip", "blank1", "wind_dir", "wind_speed", "peak_gust", "pressure", "blank2", "blank3"]
        df = pd.read_csv(url, header=None, names=columns).drop(["blank1", "blank2", "blank3"], axis=1)
        df["date"] = df["date"] + " " + df["hour"].astype(str) +":00"
        df = df.drop(["hour"], axis=1).set_index("date")
        return df
        
    def download_and_log(self, url):
        with wandb.init(project=project, job_type="load-data") as run:
            df = self.download(url)
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "met-raw", type="dataset",
                description=f"Raw meteorology data from the {self.station} weather station, split according to meteorological variables.",
                metadata={"source":url,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)
    
    def read(self, variables, artifact_name):
        artifact = wandb.Api().artifact(f"{project}/{artifact_name}:latest")
        data_folder = artifact.download()
        df = pd.DataFrame()
        for variable in variables:
            filepath = path.join(data_folder, f"{variable}.npz")
            data = np.load(filepath, allow_pickle=True)
            if df.empty:
                df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable])
            else:
                df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable]))
        return df
    
    def resample_time_and_log(self, date_index):
        variables = ["temperature", "dew_point", "humidity", "precip", "wind_dir", "wind_speed", "peak_gust", "pressure"]
        
        with wandb.init(project=project, job_type="resample-data") as run:
            artifact = run.use_artifact('met-raw:latest')
            data_folder = artifact.download()
            df = pd.DataFrame()
            for variable in variables:
                filepath = path.join(data_folder, f"{variable}.npz")
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[variable]))
            start, end = date_index.min(), date_index.max()
            df = df.loc[df.index <= end]
            df = df.loc[df.index >= start]
            # resampled_df = df.groupby(date_index[date_index.searchsorted(df.index)]).mean()
            resampled_df = df.resample("D").mean()
            columns = resampled_df.columns.to_list()
            resample_data = wandb.Artifact(
                "met-resample", type="dataset",
                description=f"Resampled meteorology data from the {self.station} weather station, split according to meteorological variables.",
                metadata={"shapes":[resampled_df[column].shape for column in columns],
                         "columns":columns})
            for column in columns:
                with resample_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=resampled_df.index, y=resampled_df[column].values)

            run.log_artifact(resample_data)
        
        return resampled_df
    
    
# Population data class

class PopData():
    def __init__(self):
        self.tmp_folder = path.join(path.abspath(""), "tmp")
        
        if not path.exists(self.tmp_folder):
            makedirs(self.tmp_folder)

    def download(self, url, verbose=False):
        self.filename = path.basename(url)
        _, self.extension = path.splitext(self.filename)
        self.filepath = path.join(self.tmp_folder, self.filename)
        
        request = requests.get(url)
        file = open(self.filepath, 'wb')
        file.write(request.content)
        file.close()
        if verbose:
            print(f"Saved to {self.filename}")
        if self.extension == ".zip":
            self.zipfiles = zpf.ZipFile(self.filepath).namelist()
            if verbose:
                print("Contains zip files:")
                [print(f"[{i}] {self.zipfiles[i]}") for i in range(len(self.zipfiles))]
        elif self.extension == ".xls":
            workbook = xlrd.open_workbook(self.filepath)
            self.sheets = workbook.sheet_names()
            if verbose:
                print(f"Contains xls sheets: {self.sheets}")
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            self.sheets = workbook.sheetnames
            if verbose:
                print(f"Contains xlsx sheets: {self.sheets}")
                
    def unzip(self, file="all", verbose=False):
        with zpf.ZipFile(self.filepath, 'r') as zip_ref:
            if file =="all":                                       # Extract all zipped files.
                zip_ref.extractall(self.tmp_folder)
            else:                                                  # Extract one specified file,
                zip_ref.extract(file, self.tmp_folder)            # then
                self.filename = path.basename(file)                # reset the file name, path and extension info.
                _, self.extension = path.splitext(self.filename)
                self.filepath = path.join(self.tmp_folder, self.filename)
             
            if verbose:
                print(f"Unzipped {file}.")
            if self.extension == ".xls":
                workbook = xlrd.open_workbook(self.filepath)
                self.sheets = workbook.sheet_names()
                if verbose:
                    print(f"Contains xls sheets: {self.sheets}")
            elif self.extension == ".xlsx":
                workbook = load_workbook(self.filepath)
                self.sheets = workbook.sheetnames
                if verbose:
                    print(f"Contains xlsx sheets: {self.sheets}")
        
    def download_and_log(self, url, region_name):
        with wandb.init(project=project, job_type="load-data") as run:
            self.download(url)
    
            df = pd.DataFrame()

            for sheet in ["Table 3", "Table 4"]:
                tmp_df = self.read_xls(sheet, verbose=False)
                tmp_df = tmp_df.loc[tmp_df[0].apply(lambda x: isinstance(x, int))].reset_index(drop=True)[[0]].rename(columns={0:"date"}).join(tmp_df.loc[tmp_df[1]=="London"].reset_index(drop=True)[[2]].rename(columns={2:"population"})).set_index("date")
                tmp_df.index = pd.to_datetime(tmp_df.index, format="%Y") + pd.tseries.offsets.DateOffset(months=6) # Set these as mid-year estimates
                if df.empty:
                    df = tmp_df.copy()
                else:
                    df = df.append(tmp_df.copy())
                    
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "population-raw", type="dataset",
                description=f"Raw annual population data for total {region_name} region. Data is extracted from source Excel file (not logged using Weights and Biases).",
                metadata={"source":url,
                         "shapes":[df[column].shape for column in columns],
                         "columns":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)
        
        [remove(path.join(self.tmp_folder, file)) for file in listdir(self.tmp_folder)]
        
    def read_csv(self, verbose=True, index_col="date", parse_dates=True):
        if verbose:
            print(f"Reading {self.filename}...")
        return pd.read_csv(self.filepath, index_col=index_col, parse_dates=parse_dates)
    
    def read_xls(self, sheet_name, verbose=False):
        if verbose:
            print(f"Reading {self.filename}...")
        if self.extension == ".xls":
            return pd.read_excel(self.filepath, sheet_name)
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            worksheet = workbook[sheet_name]
            return pd.DataFrame(worksheet.values)
        
    def read(self, artifact_name):
        artifact = wandb.Api().artifact(f"{project}/{artifact_name}:latest")
        data_folder = artifact.download()
        filepath = path.join(data_folder, f"population.npz")
        data = np.load(filepath, allow_pickle=True)
        df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=["population"])
        return df
    
    def resample_time_and_log(self, key, method):
        with wandb.init(project=project, job_type="resample-data") as run:
            artifact = run.use_artifact('population-raw:latest')
            data_folder = artifact.download()
            filepath = path.join(data_folder, f"population.npz")
            data = np.load(filepath, allow_pickle=True)
            df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"].astype(int), columns=["population"])
            
            resampled_df = df.resample(key).asfreq().interpolate(method=method)
            columns = resampled_df.columns.to_list()
            resample_data = wandb.Artifact(
                "population-resample", type="dataset",
                description=f"Resampled population data to daily resolution by {method} interpolation.",
                metadata={"shapes":[resampled_df[column].shape for column in columns],
                         "columns":columns,
                        "key":key,
                         "method":method})
            for column in columns:
                with resample_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=resampled_df.index, y=resampled_df[column].values)

            run.log_artifact(resample_data)
        
        return resampled_df  

# Disposable income data class

class IncomeData():
    def __init__(self):
        self.tmp_folder = path.join(path.abspath(""), "tmp")
        
        if not path.exists(self.tmp_folder):
            makedirs(self.tmp_folder)

    def download(self, url, verbose=False):
        self.filename = path.basename(url)
        _, self.extension = path.splitext(self.filename)
        self.filepath = path.join(self.tmp_folder, self.filename)
        
        request = requests.get(url)
        file = open(self.filepath, 'wb')
        file.write(request.content)
        file.close()
        if verbose:
            print(f"Saved to {self.filename}")
        if self.extension == ".zip":
            self.zipfiles = zpf.ZipFile(self.filepath).namelist()
            if verbose:
                print("Contains zip files:")
                [print(f"[{i}] {self.zipfiles[i]}") for i in range(len(self.zipfiles))]
        elif self.extension == ".xls":
            workbook = xlrd.open_workbook(self.filepath)
            self.sheets = workbook.sheet_names()
            if verbose:
                print(f"Contains xls sheets: {self.sheets}")
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            self.sheets = workbook.sheetnames
            if verbose:
                print(f"Contains xlsx sheets: {self.sheets}")
                
    def unzip(self, file="all", verbose=False):
        with zpf.ZipFile(self.filepath, 'r') as zip_ref:
            if file =="all":                                       # Extract all zipped files.
                zip_ref.extractall(self.tmp_folder)
            else:                                                  # Extract one specified file,
                zip_ref.extract(file, self.tmp_folder)            # then
                self.filename = path.basename(file)                # reset the file name, path and extension info.
                _, self.extension = path.splitext(self.filename)
                self.filepath = path.join(self.tmp_folder, self.filename)
             
            if verbose:
                print(f"Unzipped {file}.")
            if self.extension == ".xls":
                workbook = xlrd.open_workbook(self.filepath)
                self.sheets = workbook.sheet_names()
                if verbose:
                    print(f"Contains xls sheets: {self.sheets}")
            elif self.extension == ".xlsx":
                workbook = load_workbook(self.filepath)
                self.sheets = workbook.sheetnames
                if verbose:
                    print(f"Contains xlsx sheets: {self.sheets}")
        
    def download_and_log(self, url, region_name):
        with wandb.init(project=project, job_type="load-data") as run:
            self.download(url)
            df = self.read_xls("Table 2").transpose().set_index(0).drop(["Region", "Region name"])
            df.columns = df.loc["LAD code"].values
            df = df.drop("LAD code").dropna(axis=1)
            df.index = pd.to_datetime(np.floor(np.where(df.index.values>9999, df.index.values/10, df.index.values)).astype(int), format="%Y").rename("date")
                    
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "income-raw", type="dataset",
                description=f"Raw annual disposable income per capita data for local authorities in {region_name} region. Data is extracted from source Excel file (not logged using Weights and Biases).",
                metadata={"source":url,
                         "shapes":[df[column].shape for column in columns],
                         "LAD_codes":columns})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)
            
            metadata_df = self.read_xls("Table 2")
            metadata_df.columns = metadata_df.loc[0]
            metadata_df = pd.DataFrame(metadata_df.drop("Region", axis=1).set_index("LAD code").drop("LAD code")["Region name"]).dropna(axis=0)
            columns = metadata_df.columns.to_list()
            meta_data = wandb.Artifact(
                "income-metadata", type="dataset",
                description=f"LAD codes and corresponding local authority names for {region_name} region. Data is extracted from source Excel file (not logged using Weights and Biases).",
                metadata={"source":url,
                         "shapes":[metadata_df[column].shape for column in columns],
                         "columns":columns})
            for column in columns:
                with meta_data.new_file("LAD_codes.npz", mode="wb") as file:
                            np.savez(file, x=metadata_df.index, y=metadata_df[column].values)
            run.log_artifact(meta_data)
        
        [remove(path.join(self.tmp_folder, file)) for file in listdir(self.tmp_folder)]
        
    def read_csv(self, verbose=True, index_col="date", parse_dates=True):
        if verbose:
            print(f"Reading {self.filename}...")
        return pd.read_csv(self.filepath, index_col=index_col, parse_dates=parse_dates)
    
    def read_xls(self, sheet_name, verbose=False):
        if verbose:
            print(f"Reading {self.filename}...")
        if self.extension == ".xls":
            return pd.read_excel(self.filepath, sheet_name)
        elif self.extension == ".xlsx":
            workbook = load_workbook(self.filepath)
            worksheet = workbook[sheet_name]
            return pd.DataFrame(worksheet.values)
        
    def read(self, artifact_name):
        artifact = wandb.Api().artifact(f"{project}/{artifact_name}:latest")
        data_folder = artifact.download()
        df = pd.DataFrame()
        if artifact_name == "income-regional":
            filepath = path.join(data_folder, f"income.npz")
            data = np.load(filepath, allow_pickle=True)
            df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[f"income"])
        elif artifact_name == "income-raw" or artifact_name == "income-resample":
            for file in listdir(data_folder):
                site = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                try:
                    data = np.load(filepath, allow_pickle=True)
                    if df.empty:
                        df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                    else:
                        df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))
                except FileNotFoundError:
                    continue
        elif artifact_name == "income-metadata":
            metadata = np.load(path.join(data_folder, "LAD_codes.npz"), allow_pickle=True)
            df = pd.DataFrame(index=metadata["x"], data=metadata["y"], columns=["local_authority"])
        return df
    
    def resample_time_and_log(self, key, method):
        with wandb.init(project=project, job_type="resample-data") as run:
            artifact = run.use_artifact('income-raw:latest')
            data_folder = artifact.download()
            df = pd.DataFrame()
            for file in listdir(data_folder):
                site = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                try:
                    data = np.load(filepath, allow_pickle=True)
                    if df.empty:
                        df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"].astype(float), columns=[site])
                    else:
                        df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"].astype(float), columns=[site]))
                except FileNotFoundError:
                    continue
            
            resampled_df = df.resample(key).asfreq().interpolate(method=method)
            columns = resampled_df.columns.to_list()
            resample_data = wandb.Artifact(
                "income-resample", type="dataset",
                description=f"Resampled disposable income data to daily resolution by {method} interpolation. Local authority resolution.",
                metadata={"shapes":[resampled_df[column].shape for column in columns],
                         "columns":columns,
                        "key":key,
                         "method":method})
            for column in columns:
                with resample_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=resampled_df.index, y=resampled_df[column].values)

            run.log_artifact(resample_data)
        
        return resampled_df
    
    def regional_average_and_log(self):
        with wandb.init(project=project, job_type="regional-average-data") as run:
            artifact = run.use_artifact('income-resample:latest')
            data_folder = artifact.download()
            
            df = pd.DataFrame()
            for file in listdir(data_folder):
                site = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if df.empty:
                    df = pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site])
                else:
                    df = df.join(pd.DataFrame(index=pd.DatetimeIndex(data["x"]), data=data["y"], columns=[site]))

            df = pd.DataFrame(df.median(axis=1), columns=[f"income"])
            columns = df.columns.to_list()
            regional_data = wandb.Artifact(
                "income-regional", type="dataset",
                description=f"Regional median disposable income data at interpolated daily resolution.",
                metadata={"shapes":[df[column].shape for column in columns],
                         "columns":columns})
            for column in columns:
                with regional_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(regional_data)
        
        return df
    
    def rename_local_authority_districts(self, income_metadata_df, names_reference_list):
        income_local_authorities_list = [item for item in set(income_metadata_df.local_authority.tolist())]
        
        mismatches = []
        count = 0
        for local_authority in income_local_authorities_list:
            if local_authority not in names_reference_list:
                mismatches.append(local_authority)
            else:
                count+=1
        print(f"Pass 1:\nMatched {count} out of {len(income_local_authorities_list)} local authorities.")

        if mismatches:
            print("\nPass 2:\nSearching for matches...")
            mismatch_dict = {}
            search_letters = [local_authority[0:3] for local_authority in mismatches]
            ref_local_authorities_breakdown = [local_authority.split(" ") for local_authority in names_reference_list]
            for i in range(len(search_letters)):
                for local_authority in ref_local_authorities_breakdown:
                    for word in local_authority:
                        if search_letters[i] in word:
                            mismatch_dict.update({" ".join(local_authority): mismatches[i]})
            if mismatch_dict:
                print("Found matches (<reference_local_authority_name>: <income_local_authority_name>):")
                print(mismatch_dict)
               
                for key in mismatch_dict.keys():
                    income_metadata_df.replace({mismatch_dict[key]: key}, inplace=True)
                print("\nLocal authorities have been renamed.")
        return income_metadata_df
        
    def rename_local_authority_districts_and_log(self, reference="use_LAQN"):
        with wandb.init(project=project, job_type="rename-local-authority-data") as run:
            artifact = run.use_artifact('income-metadata:latest')
            metadata_folder = artifact.download()
            metadata = np.load(path.join(metadata_folder, "LAD_codes.npz"), allow_pickle=True)
            income_metadata_df = pd.DataFrame(index=metadata["x"], data=metadata["y"], columns=["local_authority"])

            if reference=="use_LAQN":
                meta_url = "http://api.erg.kcl.ac.uk/AirQuality/Information/MonitoringSites/GroupName=London/Json"
                sites_request = requests.get(meta_url)
                meta_df = pd.DataFrame(sites_request.json()['Sites']['Site'])
                names_reference_list = [item for item in set(meta_df["@LocalAuthorityName"].tolist())]
            else:
                names_reference_list = reference
            
            metadata_df = self.rename_local_authority_districts(income_metadata_df, names_reference_list)            
            
            columns = metadata_df.columns.to_list()
            meta_data = wandb.Artifact(
                "income-metadata", type="dataset",
                description=f"LAD codes and corresponding local authority names, after renaming using a reference.",
                metadata={"reference":reference,
                         "shapes":[metadata_df[column].shape for column in columns],
                         "columns":columns})
            for column in columns:
                with meta_data.new_file("LAD_codes.npz", mode="wb") as file:
                            np.savez(file, x=metadata_df.index, y=metadata_df[column].values)
            run.log_artifact(meta_data)
        
        return metadata_df
    
class LondonGeoData():
    def __init__(self):
        self.tmp_folder = path.join(path.abspath(""), "tmp")
        
        if not path.exists(self.tmp_folder):
            makedirs(self.tmp_folder)

    def download(self, url, verbose=False):
        
        # Get the web links for the borough & ward coordinate data files

        status, response = httplib2.Http().request(url)
        link_dict = {}

        for link in BeautifulSoup(response, parse_only=SoupStrainer('a'), features="html.parser"):
            if link.has_attr('href') and link["href"].split(".")[-1]=="zip":
                link_dict[link['href'].split("/")[-1].split(".")[0]] = f"https://data.london.gov.uk/{link['href']}"
                
        # Download the borough & ward coordinate data files and unzip them
        
        folder_path = path.join(self.tmp_folder, "i-Trees")
        if not path.exists(folder_path):
            makedirs(folder_path)
        
        for url in (progress_bar := tqdm(link_dict.values())):
            progress_bar.set_description(f"Downloading {path.basename(url)}")
            request = requests.get(url)
            filepath = path.join(folder_path, path.basename(url))
            file = open(filepath, 'wb')
            file.write(request.content)
            file.close()
            zpf.ZipFile(filepath, 'r').extractall(path.join(self.tmp_folder, "i-Trees"))
        
        return link_dict
            
    def process(self, link_dict):
        # Compile a geopandas dataframe of the london wards coordinates

        london_wards_gdf = gpd.GeoDataFrame()

        for borough in (progress_bar := tqdm(link_dict.keys())):
            progress_bar.set_description(f"Processing {borough}")
            borough_folder = path.join(self.tmp_folder, "i-Trees", borough)
            shapefiles = [file for file in listdir(borough_folder) if file.split(".")[-1]=="shp"]
            for shapefile in shapefiles:
                gdf = gpd.read_file(path.join(borough_folder, shapefile))
                if london_wards_gdf.empty:
                    london_wards_gdf = gdf
                else:
                    london_wards_gdf = pd.concat([london_wards_gdf, gdf])
        
        return london_wards_gdf
    
    def download_and_log(self, local_authority_scale=True, ward_scale=False):
        with wandb.init(project=project, job_type="load-data") as run:
            url = "https://data.london.gov.uk/dataset/i-trees-canopy-ward-data"
            link_dict = self.download(url)
            london_wards_gdf = self.process(link_dict)

            # Aggregate to local authority spatial scale (London boroughs)
            london_authorities_gdf = london_wards_gdf[["BOROUGH", "geometry"]].dissolve(by="BOROUGH").reset_index()
            london_authorities_gdf.rename(columns={"BOROUGH":"local_authority"}, inplace=True)

            # Log the London wards shapefile with wandb
            df = london_wards_gdf.copy()
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "london-wards-raw", type="dataset",
                description=f"Raw shapefile for London wards. Data is extracted from i-Trees project on London Datastore (not logged using Weights and Biases).",
                metadata={"source":url,
                         "shapes":[df[column].shape for column in columns]})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)

            # Log the London local authorities shapefile with wandb
            df = london_authorities_gdf.copy()
            columns = df.columns.to_list()

            raw_data = wandb.Artifact(
                "london-local-authorities-raw", type="dataset",
                description=f"Raw shapefile for London local authorities. Data is extracted & processed from i-Trees project on London Datastore (not logged using Weights and Biases).",
                metadata={"source":url,
                         "shapes":[df[column].shape for column in columns]})

            for column in columns:
                with raw_data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=df.index, y=df[column].values)

            run.log_artifact(raw_data)

            rmtree(path.join(self.tmp_folder, "i-Trees"))
            print("Done!")
            
            if local_authority_scale:
                return london_authorities_gdf
            elif ward_scale:
                return london_wards_gdf
        
    def read(self, artifact_name):
        artifact = wandb.Api().artifact(f"{project}/{artifact_name}:latest")
        data_folder = artifact.download()

        if artifact_name == "edge-pairs-array":
            filepath = path.join(data_folder, "edge_pairs.npz")
            data = np.load(filepath, allow_pickle=True)
            array = np.array([data["x"], data["y"]])
            return array
        else:
            gdf = gpd.GeoDataFrame()
            for file in listdir(data_folder):
                column = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if gdf.empty:
                    gdf = gpd.GeoDataFrame(index=data["x"], data=data["y"], columns=[column])
                else:
                    gdf = gdf.join(gpd.GeoDataFrame(index=data["x"], data=data["y"], columns=[column]))
        return gdf
    
    def rename_local_authority_districts(self, df, names_reference_list):
        local_authorities_to_rename = [item for item in set(df.local_authority.tolist())]
        
        mismatches = []
        count = 0
        for local_authority in local_authorities_to_rename:
            if local_authority not in names_reference_list:
                mismatches.append(local_authority)
            else:
                count+=1
        print(f"Pass 1:\nMatched {count} out of {len(local_authorities_to_rename)} local authorities.")

        if mismatches:
            print("\nPass 2:\nSearching for matches...")
            mismatch_dict = {}
            search_letters = [local_authority[0:3] for local_authority in mismatches]
            ref_local_authorities_breakdown = [local_authority.split(" ") for local_authority in names_reference_list]
            for i in range(len(search_letters)):
                for local_authority in ref_local_authorities_breakdown:
                    for word in local_authority:
                        if search_letters[i] in word:
                            mismatch_dict.update({" ".join(local_authority): mismatches[i]})
            if mismatch_dict:
                print("Found matches (<reference_local_authority_name>: <local_authority_to_rename>):")
                print(mismatch_dict)
               
                for key in mismatch_dict.keys():
                    df.replace({mismatch_dict[key]: key}, inplace=True)
                print("\nLocal authorities have been renamed.")
        return df
    
    def rename_local_authority_districts_and_log(self, names_reference_list, reference):
        with wandb.init(project=project, job_type="rename-local-authority-data") as run:
            artifact = run.use_artifact(f'london-local-authorities-raw:latest')
            data_folder = artifact.download()
            gdf = gpd.GeoDataFrame()

            for file in listdir(data_folder):
                column = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if gdf.empty:
                    gdf = gpd.GeoDataFrame(index=data["x"], data=data["y"], columns=[column])
                else:
                    gdf = gdf.join(gpd.GeoDataFrame(index=data["x"], data=data["y"], columns=[column]))
                    
            gdf = self.rename_local_authority_districts(gdf, names_reference_list)
            
            # Log the renamed London local authorities shapefile with wandb
            columns = gdf.columns.to_list()

            data = wandb.Artifact(
                "london-local-authorities-renamed", type="dataset",
                description=f"Processed shapefile for London local authorities, renamed using a reference.",
                metadata={"reference":reference,
                          "columns":columns,
                         "shapes":[gdf[column].shape for column in columns]})

            for column in columns:
                with data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=gdf.index, y=gdf[column].values)

            run.log_artifact(data)
            return gdf
        
    def get_local_authority_neighbour_edges(self, london_authorities_gdf):
        london_authorities_gdf["neighbours"] = None
        edge_pairs = []

        for index, row in london_authorities_gdf.iterrows():
            # get adjoining local authorities
            neighbours = london_authorities_gdf[london_authorities_gdf.geometry.touches(row.geometry)].local_authority.tolist()
            neighbours_index = london_authorities_gdf[london_authorities_gdf.geometry.touches(row.geometry)].index.tolist()

            # add names of neighbours as row value
            london_authorities_gdf.at[index, "neighbours"] = neighbours

            # add the neighbouring pair of indices to the edge_pairs list
            for neighbour in neighbours_index:
                edge_pairs.append([index, neighbour])
        
        edge_pairs = np.array(edge_pairs).transpose()
        return london_authorities_gdf, edge_pairs
    
    def get_local_authority_neighbour_edges_and_log(self):
        with wandb.init(project=project, job_type="get-local-authority-neighbour-edge-data") as run:
            artifact = run.use_artifact(f'london-local-authorities-renamed:latest')
            data_folder = artifact.download()
            gdf = gpd.GeoDataFrame()

            for file in listdir(data_folder):
                column = file.replace(".npz", "")
                filepath = path.join(data_folder, file)
                data = np.load(filepath, allow_pickle=True)
                if gdf.empty:
                    gdf = gpd.GeoDataFrame(index=data["x"], data=data["y"], columns=[column])
                else:
                    gdf = gdf.join(gpd.GeoDataFrame(index=data["x"], data=data["y"], columns=[column]))
                    
            gdf, edge_array = self.get_local_authority_neighbour_edges(gdf)

            
            # Log the updated London local authorities shapefile with wandb
            columns = gdf.columns.to_list()

            data = wandb.Artifact(
                "london-local-authorities-neighbours", type="dataset",
                description=f"Processed shapefile for London local authorities, with neighbouring local authorities identified.",
                metadata={"columns":columns,
                         "shapes":[gdf[column].shape for column in columns]})

            for column in columns:
                with data.new_file(column + ".npz", mode="wb") as file:
                        np.savez(file, x=gdf.index, y=gdf[column].values)

            run.log_artifact(data)
            
            # Log the edge pairs numpy array with wandb
            data = wandb.Artifact(
                "edge-pairs-array", type="dataset",
                description=f"Edge pairs array for neighbouring London local authorities.",
                metadata={"shape":edge_array.shape})

            with data.new_file("edge_pairs" + ".npz", mode="wb") as file:
                np.savez(file, x=edge_array[0], y=edge_array[1])

            run.log_artifact(data)
            
            return gdf, edge_array